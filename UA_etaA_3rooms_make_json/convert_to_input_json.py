import openpyxl
import json


# # Convert EXCEL sheet to json format

def count_number_in_id_row(sheet):
    id_all = [row[1].value for row in sheet.rows][1:]
    return len(id_all) - (id_all).count(None)

def make_dictionary_of_layer(row):
    n = int(row[2].value)
    layer = [
        {
            "name": row[3+3*i].value,
            "thermal_resistance": float(row[4+3*i].value),
            "thermal_capacity": float(row[5+3*i].value)
        } for i in range(n)
    ]
    # Tuple(layer_list, reversed_layer_list)
    return layer, layer[::-1]


def get_h_c(direction):
    if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
        return 2.5
    elif direction == 'bottom':
        return 0.7
    elif direction == 'top':
        return 5.0
    elif direction == 'horizontal':
        return (2.5, 2.5)
    elif direction == 'upward':
        return (5.0, 0.7)
    elif direction == 'downward':
        return (0.7, 5.0)
    else:
        raise Exception()


def get_outside_heat_transfer_resistance(direction):
    if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
        return 0.04
    elif direction == 'bottom':
        return 0.15
    elif direction == 'top':
        return 0.09
    elif direction == 'horizontal':
        return (0.11, 0.11)
    elif direction == 'upward':
        return (0.09, 0.15)
    elif direction == 'downward':
        return (0.15, 0.09)
    else:
        raise Exception()


def get_solar_shading(exist: bool, depth=None, d_h=None, d_e=None):
    if exist:
        return {
            "existence": True,
            "input_method": "simple",
            "depth": float(depth),
            "d_h": float(d_h),
            "d_e": float(d_e)
        }
    else:
        return {
            "existence": False
        }


def get_is_floor(direction):
    if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se', 'top']:
        return False
    elif direction == 'bottom':
        return True
    elif direction == 'horizontal':
        return (False, False)
    elif direction == 'upward':
        return (False, True)
    elif direction == 'downward':
        return (True, False)
    else:
        raise Exception()


def convert_excel_to_json(excel_file: str) -> dict:

    book = openpyxl.load_workbook(excel_file, data_only=True)

    sheet_common = book['common']
    sheet_building = book['building']
    sheet_rooms = book['rooms']
    sheet_external_general_parts = book['external_general_parts']
    sheet_external_opaque_parts = book['external_opaque_parts']
    sheet_external_transparent_parts = book['external_transparent_parts']
    sheet_internals = book['internals']
    sheet_grounds = book['grounds']
    sheet_layers = book['layers']


    n_rooms = count_number_in_id_row(sheet=sheet_rooms)
    n_rooms

    n_external_general_parts = count_number_in_id_row(sheet=sheet_external_general_parts)
    n_external_general_parts

    n_external_opaque_parts = count_number_in_id_row(sheet=sheet_external_opaque_parts)
    n_external_opaque_parts

    n_external_transparent_parts = count_number_in_id_row(sheet=sheet_external_transparent_parts)
    n_external_transparent_parts

    n_internals = count_number_in_id_row(sheet=sheet_internals)
    n_internals

    n_grounds = count_number_in_id_row(sheet=sheet_grounds)
    n_grounds

    n_layers = count_number_in_id_row(sheet=sheet_layers)
    n_layers


    def get_layers(layer_name, is_reverse=False):
        # use variable 'layers_master' as global variable
        layers = list(filter(lambda d: d['name'] == layer_name , layers_master))
        if len(layers) > 1:
            raise Exception("Match over one layer.")
        if len(layers) == 0:
            raise Exception("Can't find the layer", layer_name)
        if is_reverse:
            return layers[0]['reversed_layers']
        else:
            return layers[0]['layers']
    
    layers_master = [
        {
            "name": row[1].value,
            "layers": make_dictionary_of_layer(row)[0],
            "reversed_layers": make_dictionary_of_layer(row)[1]
        } for row in sheet_layers.iter_rows(min_row=2, max_row=n_layers+1)
    ]

    common = {
        'ac_method': sheet_common.cell(column=2, row=2).value
    }

    building = {
        "infiltration": {
            "method": "balance_residential",
            "c_value_estimate": "specify",
            "story": int(sheet_building.cell(column=2, row=2).value),
            "c_value": float(sheet_building.cell(column=3, row=2).value),
            "inside_pressure": sheet_building.cell(column=4, row=2).value
        }
    }

    rooms = [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value,
            "floor_area": float(row[4].value),
            "volume": float(row[5].value),
            "ventilation": {
                "natural": float(row[6].value)
            },
            "furniture": {
                "input_method": "specify",
                "heat_capacity": float(row[8].value),
                "heat_cond": 0.00022 * float(row[8].value),
                "moisture_capacity": 0.0,
                "moisture_cond": 0.9
            },
            "schedule": {
                "name": row[7].value
            }
        } for row in sheet_rooms.iter_rows(min_row=2, max_row=n_rooms+1)
    ]

    external_general_parts =  [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value,
            "connected_room_id": int(row[4].value),
            "boundary_type": "external_general_part",
            "area": float(row[5].value),
            "h_c": get_h_c(direction=row[8].value),
            "is_solar_absorbed_inside": bool(row[6].value),
            "is_floor": bool(row[6].value),
            "layers": get_layers(layer_name=row[7].value),
            "solar_shading_part": {"existence": False},
            "is_sun_striked_outside": True,
            "direction": row[8].value,
            "outside_emissivity": 0.9,
            "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[8].value),
            "outside_solar_absorption": 0.8,
            "temp_dif_coef": float(row[9].value)
        } for row in sheet_external_general_parts.iter_rows(min_row=2, max_row=n_external_general_parts+1) if float(row[5].value) > 0.0
    ]


    external_opaque_parts =  [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value,
            "connected_room_id": int(row[4].value),
            "boundary_type": "external_opaque_part",
            "area": float(row[5].value),
            "h_c": get_h_c(direction=row[7].value),
            "is_solar_absorbed_inside": False,
            "is_floor": False,
            "solar_shading_part": {"existence": False},
            "is_sun_striked_outside": True,
            "direction": row[7].value,
            "outside_emissivity": 0.9,
            "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[7].value),
            "u_value": float(row[6].value),
            "inside_heat_transfer_resistance": 0.11,
            "outside_solar_absorption": 0.8,
            "temp_dif_coef": 1.0
        } for row in sheet_external_opaque_parts.iter_rows(min_row=2, max_row=n_external_opaque_parts+1) if float(row[5].value) > 0.0
    ]

    external_transparent_parts =  [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value,
            "connected_room_id": int(row[4].value),
            "boundary_type": "external_transparent_part",
            "area": float(row[5].value),
            "h_c": get_h_c(direction=row[10].value),
            "is_solar_absorbed_inside": False,
            "is_floor": False,
            "solar_shading_part": get_solar_shading(exist=bool(row[11].value), depth=row[12].value, d_h=row[13].value, d_e=row[14].value),
            "is_sun_striked_outside": True,
            "direction": row[10].value,
            "outside_emissivity": 0.9,
            "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[10].value),
            "u_value": float(row[6].value),
            "inside_heat_transfer_resistance": 0.11,
            "eta_value": float(row[7].value),
            "incident_angle_characteristics": row[8].value,
            "glass_area_ratio": float(row[9].value),
            "temp_dif_coef": 1.0
        } for row in sheet_external_transparent_parts.iter_rows(min_row=2, max_row=n_external_transparent_parts+1) if float(row[5].value) > 0.0
    ]

    internals_2d =  [
        [
            {
                "id": row[1].value,
                "name": row[3].value,
                "sub_name": row[5].value,
                "connected_room_id": int(row[7].value),
                "boundary_type": "internal",
                "area": float(row[9].value),
                "h_c": get_h_c(direction=row[11].value)[0],
                "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[0],
                "is_floor": get_is_floor(direction=row[11].value)[0],
                "layers": get_layers(layer_name=row[10].value, is_reverse=False),
                "solar_shading_part": get_solar_shading(exist=False),
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[11].value)[0],
                "rear_surface_boundary_id": row[2].value
            },
            {
                "id": row[2].value,
                "name": row[4].value,
                "sub_name": row[6].value,
                "connected_room_id": int(row[8].value),
                "boundary_type": "internal",
                "area": float(row[9].value),
                "h_c": get_h_c(direction=row[11].value)[1],
                "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[1],
                "is_floor": get_is_floor(direction=row[11].value)[1],
                "layers": get_layers(layer_name=row[10].value, is_reverse=True),
                "solar_shading_part": get_solar_shading(exist=False),
                "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[11].value)[1],
                "rear_surface_boundary_id": row[1].value
            }
        ] for row in sheet_internals.iter_rows(min_row=2, max_row=n_internals+1) if float(row[9].value) > 0.0
    ]
    # flatten
    internals = sum(internals_2d, [])

    grounds =  [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value,
            "connected_room_id": int(row[4].value),
            "boundary_type": "ground",
            "area": float(row[5].value),
            "is_solar_absorbed_inside": bool(row[7].value),
            "is_floor": True,
            "h_c": get_h_c(direction='bottom'),
            "layers": get_layers(layer_name=row[6].value)
        } for row in sheet_grounds.iter_rows(min_row=2, max_row=n_grounds+1) if float(row[5].value) > 0.0
    ]

    complete_list = {
        "common":common,
        "building": building,
        "rooms": rooms,
        "boundaries": external_general_parts + external_opaque_parts + external_transparent_parts + internals + grounds
    }

    return complete_list


if __name__ == '__main__':

    d = convert_excel_to_json('continuous_calc_input_excel.xlsx')

